from flask import Flask, request, send_file, jsonify
import pandas as pd
from urllib.parse import quote as url_quote
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os

app = Flask(__name__)

# Enable CORS for frontend communication
from flask_cors import CORS
CORS(app, resources={r"/*": {"origins": "*"}})

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def process_excel(file1_path, file2_path, output_path):
    sheet1 = pd.read_excel(file1_path, header=None)
    sheet2 = pd.read_excel(file2_path, header=None)

    output_sheet = sheet1.iloc[:2, :].copy()
    compare_data_1 = sheet1.iloc[2:, 1:].reset_index(drop=True)
    compare_data_2 = sheet2.iloc[2:, 1:].reset_index(drop=True)

    mismatched_rows = compare_data_1[~compare_data_1.apply(tuple, axis=1).isin(compare_data_2.apply(tuple, axis=1))]
    mismatched_rows.insert(0, 0, range(1, len(mismatched_rows) + 1))

    output_sheet = pd.concat([output_sheet, mismatched_rows], ignore_index=True)
    output_sheet.to_excel(output_path, index=False, header=False)

    # Apply Formatting
    wb = load_workbook(output_path)
    ws = wb.active

    styles = {
        "bold_white": Font(color='FFFFFF', size=10, name='Arial', bold=True),
        "regular_black": Font(color='000000', size=10, name='Arial', bold=False),
        "solid_fill": PatternFill(start_color='FF365838', end_color='FF365838', fill_type='solid'),
        "border_white": Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        ),
        "align_center": Alignment(horizontal='center', vertical='center'),
        "align_left": Alignment(horizontal='left', vertical='center')
    }

    for row in range(1, 3):
        for col in range(1, 18):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["solid_fill"]
            cell.font = styles["bold_white"]
            cell.alignment = styles["align_left"] if col == 1 and row == 1 else styles["align_center"]
            if row == 2:
                cell.border = styles["border_white"]

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=17):
        for cell in row:
            cell.font = styles["regular_black"]
            cell.alignment = styles["align_center"]

    ws.column_dimensions['A'].width = 10
    for col in range(2, 18):
        col_letter = ws.cell(row=1, column=col).column_letter
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in ws[col_letter])
        ws.column_dimensions[col_letter].width = max_length + 2  

    wb.save(output_path)

@app.route('/upload', methods=['POST'])
def upload_files():
    if 'file1' not in request.files or 'file2' not in request.files:
        return jsonify({"error": "Both files are required"}), 400

    file1 = request.files['file1']
    file2 = request.files['file2']

    file1_path = os.path.join(UPLOAD_FOLDER, file1.filename)
    file2_path = os.path.join(UPLOAD_FOLDER, file2.filename)
    output_path = os.path.join(OUTPUT_FOLDER, "output_sheet3.xlsx")

    file1.save(file1_path)
    file2.save(file2_path)

    process_excel(file1_path, file2_path, output_path)

    return jsonify({"message": "Convert chesi Dhobba mawa", "download_url": "/download"}), 200

@app.route('/download', methods=['GET'])
def download_file():
    output_path = os.path.join(OUTPUT_FOLDER, "output_sheet3.xlsx")
    return send_file(output_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
